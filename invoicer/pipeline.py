"""
Invoice fill + export (used by the Django invoicer app).

Templates live in ``assets/invoice/`` (see project ``INVOICE_ASSET_DIR`` setting).
"""
from __future__ import annotations

import calendar
import os
import shutil
import subprocess
import sys
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

_PACKAGE_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _PACKAGE_DIR.parent
# Default when ``asset_dir`` is not passed (e.g. tests); Django always sets INVOICE_ASSET_DIR.
INVOICE_ASSET_DIR = _PROJECT_ROOT / "assets" / "invoice"


class InvoiceError(Exception):
    """Raised when invoice generation fails (templates, PDF, validation)."""

DATA_FIRST_ROW = 17
DATA_LAST_ROW = 25
SUM_ROW = 26
DATE_LABEL_CELL = "B29"
TABLE_HEADER_ROW = 16
ALIGN_COLS = ("B", "C", "E", "F")

CELL_CENTER = Alignment(horizontal="center", vertical="center")
CELL_LEFT = Alignment(horizontal="left", vertical="center")

INVOICE_WEEK_1_MONDAY = date(2025, 6, 30)

XLSX_NAMES = ("Template.xlsx", "template.xlsx")
NUMBERS_NAMES = (
    "invoice_template.numbers",
    "Template.numbers",
    "template.numbers",
    "invoice template.numbers",
    "Invoice Template.numbers",
)


def find_soffice() -> str | None:
    mac = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    if mac.is_file():
        return str(mac)
    win = Path(r"C:\Program Files\LibreOffice\program\soffice.exe")
    if win.is_file():
        return str(win)
    for name in ("soffice", "libreoffice"):
        p = shutil.which(name)
        if p:
            return p
    return None


def _convert_xlsx_to_pdf_libreoffice(xlsx_path: Path, pdf_path: Path) -> bool:
    soffice = find_soffice()
    if not soffice:
        return False
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        cmd = [
            soffice,
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--convert-to",
            "pdf",
            "--outdir",
            str(tmp),
            str(xlsx_path.resolve()),
        ]
        r = subprocess.run(cmd, capture_output=True, text=True)
        if r.returncode != 0:
            msg = r.stderr or r.stdout or "LibreOffice failed"
            raise InvoiceError(msg)
        produced = tmp / (xlsx_path.stem + ".pdf")
        if not produced.is_file():
            raise InvoiceError(f"Expected PDF not found: {produced}")
        shutil.copy2(produced, pdf_path)
    return True


def export_xlsx_to_pdf_via_numbers(xlsx_path: Path, pdf_path: Path) -> None:
    """Open xlsx in Numbers and export PDF (macOS only)."""
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    if pdf_path.is_file():
        pdf_path.unlink()
    src = str(xlsx_path.resolve())
    dst = str(pdf_path.resolve())
    src_esc = src.replace("\\", "\\\\").replace('"', '\\"')
    dst_esc = dst.replace("\\", "\\\\").replace('"', '\\"')
    script = f"""
tell application "Numbers"
    activate
    open POSIX file "{src_esc}"
    delay 3
    export front document to POSIX file "{dst_esc}" as PDF
    close front document saving no
end tell
"""
    r = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True,
        text=True,
    )
    if r.returncode != 0 or not pdf_path.is_file():
        msg = [
            "Numbers could not export the PDF (LibreOffice is not installed).",
            "Grant Terminal/Cursor automation access for Numbers in System Settings → Privacy & Security → Automation,",
            "or install LibreOffice for headless PDF conversion.",
        ]
        if r.stderr:
            msg.append(r.stderr.strip())
        if r.stdout:
            msg.append(r.stdout.strip())
        raise RuntimeError("\n".join(m for m in msg if m))


def convert_xlsx_to_pdf(xlsx_path: Path, pdf_path: Path) -> None:
    if _convert_xlsx_to_pdf_libreoffice(xlsx_path, pdf_path):
        return
    if sys.platform == "darwin":
        try:
            export_xlsx_to_pdf_via_numbers(xlsx_path, pdf_path)
            return
        except RuntimeError as exc:
            sys.stderr.write(f"{exc}\n")
    raise InvoiceError(
        "Could not create PDF. Install LibreOffice or fix Numbers automation on macOS. "
        "See https://www.libreoffice.org/download/download/"
    )


def resolve_numbers_file(asset_dir: Path | None = None) -> Path | None:
    root = asset_dir or INVOICE_ASSET_DIR
    for name in NUMBERS_NAMES:
        p = root / name
        if p.is_file():
            return p
    numbers = sorted(root.glob("*.numbers"))
    if len(numbers) == 1:
        return numbers[0]
    if len(numbers) > 1:
        raise InvoiceError(
            "Multiple .numbers files found; keep one or rename to Template.numbers. "
            f"Found: {[f.name for f in numbers]}"
        )
    return None


def resolve_xlsx_template(asset_dir: Path | None = None) -> Path | None:
    root = asset_dir or INVOICE_ASSET_DIR
    for name in XLSX_NAMES:
        p = root / name
        if p.is_file():
            return p
    return None


def _posix_path_for_as(path: Path) -> str:
    return str(path.resolve()).replace("\\", "\\\\").replace('"', '\\"')


def _as_cell_string(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')


def format_date_dmy(d: date | datetime) -> str:
    """Day/month/year text, no time (e.g. 15/07/2025)."""
    if isinstance(d, datetime):
        d = d.date()
    return d.strftime("%d/%m/%Y")


def apply_table_center_alignment_openpyxl(ws) -> None:
    for row in range(TABLE_HEADER_ROW, SUM_ROW + 1):
        for col in ALIGN_COLS:
            ws[f"{col}{row}"].alignment = CELL_CENTER
    ws["F3"].alignment = CELL_CENTER


def monday_of_week_au(day: date) -> date:
    """Return the Monday starting the Australian (ISO) week that contains ``day``."""
    return day - timedelta(days=day.weekday())


def output_pdf_path_for_week(week_monday: date, output_dir: Path) -> Path:
    """
    PDF name from ISO week of the invoice Monday: Week_{n}_{year}.pdf
    e.g. Week_1_2026.pdf for ISO week 1 of 2026.
    """
    iso_year, iso_week, _ = week_monday.isocalendar()
    return output_dir / f"Week_{iso_week}_{iso_year}.pdf"


def get_invoice_number(week_reference_date: date | datetime) -> int:
    """
    Invoice numbers count Monday-based weeks from INVOICE_WEEK_1_MONDAY:
    1 = week of 30 Jun–6 Jul 2025, 2 = week starting 7 Jul 2025, etc.
    """
    d = week_reference_date.date() if isinstance(week_reference_date, datetime) else week_reference_date
    ref_monday = monday_of_week_au(d)
    delta_weeks = (ref_monday - INVOICE_WEEK_1_MONDAY).days // 7
    return 1 + delta_weeks


def get_date(rows: list[tuple[datetime, int]]) -> str:
    """
    From ``rows``: Australian Mon–Sun week of the earliest date → week-end Sunday → +14 days
    (next-next Sunday), as DD/MM/YYYY.
    """
    earliest = min(dt.date() for dt, _ in rows)
    week_monday = monday_of_week_au(earliest)
    week_end_sunday = week_monday + timedelta(days=6)
    next_next_sunday = week_end_sunday + timedelta(days=14)
    return format_date_dmy(next_next_sunday)


def fill_invoice_via_numbers(
    numbers_path: Path,
    rows_data: list[tuple[datetime, int]],
    invoice_number: int,
    output_pdf: Path,
    output_xlsx: Path,
) -> None:
    """
    Open the .numbers template in Numbers, set cells (preserves images), export
    output.xlsx and ``output_pdf``. Does not use openpyxl on the template.
    """
    if sys.platform != "darwin":
        raise RuntimeError("Numbers automation requires macOS.")

    sheet_n = int(os.environ.get("NUMBERS_SHEET", "1"))
    table_n = int(os.environ.get("NUMBERS_TABLE", "1"))
    src_esc = _posix_path_for_as(numbers_path)
    xlsx_esc = _posix_path_for_as(output_xlsx)
    pdf_esc = _posix_path_for_as(output_pdf)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    if output_xlsx.is_file():
        output_xlsx.unlink()
    if output_pdf.is_file():
        output_pdf.unlink()

    body: list[str] = [
        'tell application "Numbers"',
        "    activate",
        f'    open POSIX file "{src_esc}"',
        "    delay 3",
        "    tell front document",
        f"        tell sheet {sheet_n}",
        f"            tell table {table_n}",
    ]
    for r in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
        for col in ("B", "C", "E", "F"):
            body.append(f'                set value of cell "{col}{r}" to ""')
    for i, (dt, parcels) in enumerate(rows_data):
        r = DATA_FIRST_ROW + i
        dmy = _as_cell_string(format_date_dmy(dt))
        body.append(f'                set value of cell "B{r}" to "{dmy}"')
        wd = _as_cell_string(weekday_english(dt))
        body.append(f'                set value of cell "C{r}" to "{wd}"')
        body.append(f'                set value of cell "E{r}" to {parcels}')
        body.append(f'                set value of cell "F{r}" to "=E{r}*$G$15"')
    body.append('                set value of cell "E26" to "=SUM(E17:E25)"')
    body.append('                set value of cell "F26" to "=SUM(F17:F25)"')
    body.append(f'                set value of cell "F3" to {invoice_number}')
    b29_text = _as_cell_string(f"DATE : {get_date(rows_data)}")
    body.append(f'                set value of cell "{DATE_LABEL_CELL}" to "{b29_text}"')
    body.append("                try")
    body.append('                    set alignment of cell "F3" to center')
    body.append(f"                    repeat with rowNum from {TABLE_HEADER_ROW} to {SUM_ROW}")
    body.append('                        repeat with colLetter in {"B", "C", "E", "F"}')
    body.append(
        '                            set alignment of cell ((colLetter as text) & rowNum) to center'
    )
    body.append("                        end repeat")
    body.append("                    end repeat")
    body.append("                end try")
    body.extend(
        [
            "            end tell",
            "        end tell",
            f'        export to POSIX file "{xlsx_esc}" as Microsoft Excel',
            f'        export to POSIX file "{pdf_esc}" as PDF',
            "    end tell",
            "    close front document saving no",
            "end tell",
        ]
    )
    script = "\n".join(body)
    r = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True,
        text=True,
    )
    if r.returncode != 0 or not output_pdf.is_file():
        msg = [
            "Numbers fill or export failed.",
            "Check NUMBERS_SHEET / NUMBERS_TABLE if cells land in the wrong place.",
            "Allow Terminal/Cursor to control Numbers (System Settings → Privacy → Automation).",
        ]
        if r.stderr:
            msg.append(r.stderr.strip())
        if r.stdout:
            msg.append(r.stdout.strip())
        raise RuntimeError("\n".join(m for m in msg if m))


def weekday_english(dt: datetime) -> str:
    return calendar.day_name[dt.weekday()]


def fill_workbook(
    wb_path: Path,
    rows_data: list[tuple[datetime, int]],
    invoice_number: int,
    output_xlsx: Path,
) -> None:
    wb = load_workbook(wb_path)
    ws = wb.active

    ws["F3"] = invoice_number

    for r in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
        ws[f"B{r}"] = None
        ws[f"C{r}"] = None
        ws[f"E{r}"] = None
        ws[f"F{r}"] = None

    for i, (dt, parcels) in enumerate(rows_data):
        r = DATA_FIRST_ROW + i
        cell_b = ws[f"B{r}"]
        cell_b.value = format_date_dmy(dt)
        cell_b.number_format = "@"
        ws[f"C{r}"] = weekday_english(dt)
        ws[f"E{r}"] = parcels
        ws[f"F{r}"] = f"=E{r}*$G$15"

    ws["E26"] = "=SUM(E17:E25)"
    ws["F26"] = "=SUM(F17:F25)"

    apply_table_center_alignment_openpyxl(ws)

    c29 = ws[DATE_LABEL_CELL]
    c29.value = f"DATE : {get_date(rows_data)}"
    c29.font = Font(bold=True)
    c29.alignment = CELL_LEFT
    c29.number_format = "@"

    wb.save(output_xlsx)


def run_invoice_pipeline(
    rows: list[tuple[datetime, int]],
    *,
    output_dir: Path,
    asset_dir: Path | None = None,
) -> tuple[Path, Path, int]:
    """
    Build invoice XLSX + week-named PDF. Returns (xlsx_path, pdf_path, invoice_number).
    Session ``output_dir`` holds ``output.xlsx`` and ``Week_*_*.pdf`` (regenerated each run).
    """
    if not rows:
        raise InvoiceError("Add at least one delivery line.")
    max_lines = DATA_LAST_ROW - DATA_FIRST_ROW + 1
    if len(rows) > max_lines:
        raise InvoiceError(f"Too many rows (max {max_lines}).")

    ad = asset_dir or INVOICE_ASSET_DIR
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    ref_date = min(dt.date() for dt, _ in rows)
    invoice_number = get_invoice_number(ref_date)
    week_monday = monday_of_week_au(ref_date)
    out_pdf = output_pdf_path_for_week(week_monday, output_dir)
    out_xlsx = output_dir / "output.xlsx"

    numbers = resolve_numbers_file(ad)
    xlsx_tpl = resolve_xlsx_template(ad)

    if sys.platform == "darwin" and numbers is not None:
        fill_invoice_via_numbers(numbers, rows, invoice_number, out_pdf, out_xlsx)
        return out_xlsx, out_pdf, invoice_number

    if xlsx_tpl is not None:
        fill_workbook(xlsx_tpl, rows, invoice_number, out_xlsx)
        convert_xlsx_to_pdf(out_xlsx, out_pdf)
        return out_xlsx, out_pdf, invoice_number

    if numbers is not None:
        raise InvoiceError(
            "Found a .numbers template but Numbers automation only works on macOS. "
            f"Export to {XLSX_NAMES[0]} or run on a Mac."
        )

    raise InvoiceError(
        f"No invoice template in {ad}. Add a .numbers file or {XLSX_NAMES[0]}."
    )
