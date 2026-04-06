#!/usr/bin/env python3
"""
Generate assets/invoice/Template.xlsx for Linux (openpyxl + LibreOffice PDF).

For a PDF that matches your Numbers design, export from Numbers:
  File → Export To → Excel… → save as Template.xlsx in assets/invoice/
and replace this file. This script only builds a structured fallback with the
same cell map as invoicer/pipeline.py (rows 17–25, F3, G15, B29, sums E26/F26).
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins

OUT = Path(__file__).resolve().parent.parent / "assets" / "invoice" / "Template.xlsx"

THIN = Side(style="thin", color="FF334155")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill(start_color="FFF1F5F9", end_color="FFF1F5F9", fill_type="solid")


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws.merge_cells("A1:G1")
    ws["A1"] = "TAX INVOICE"
    ws["A1"].font = Font(size=20, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Delivery lines and totals follow the table below."
    ws["A3"].font = Font(size=10, italic=True, color="FF64748B")
    ws.merge_cells("A3:D4")

    ws["E3"] = "Invoice #"
    ws["E3"].font = Font(bold=True, size=11)
    ws["E3"].alignment = Alignment(horizontal="right", vertical="center")

    ws["G14"] = "Rate / parcel"
    ws["G14"].font = Font(size=10, bold=True)
    ws["G14"].alignment = Alignment(horizontal="right", vertical="center")
    ws["G15"] = 1.0
    ws["G15"].number_format = "#,##0.00"
    ws["G15"].alignment = Alignment(horizontal="right", vertical="center")

    headers = (
        ("B16", "Date"),
        ("C16", "Day"),
        ("D16", ""),
        ("E16", "Parcels"),
        ("F16", "Amount"),
    )
    for addr, label in headers:
        cell = ws[addr]
        cell.value = label
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = HDR_FILL
        cell.border = BOX

    for r in range(17, 26):
        for col in range(2, 7):
            ws.cell(row=r, column=col).border = BOX
            ws.cell(row=r, column=col).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    for col, letter in enumerate("BCDEF", start=2):
        ws.cell(row=26, column=col).border = BOX

    ws["D26"] = "TOTAL"
    ws["D26"].font = Font(bold=True, size=11)
    ws["D26"].alignment = Alignment(horizontal="right", vertical="center")
    ws["E26"].alignment = Alignment(horizontal="center", vertical="center")
    ws["F26"].alignment = Alignment(horizontal="center", vertical="center")

    ws["B29"] = "DATE : (filled when generated)"
    ws["B29"].font = Font(bold=True, size=11)
    ws["B29"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[29].height = 22

    for col, w in ("A", 4), ("B", 14), ("C", 14), ("D", 6), ("E", 11), ("F", 14), ("G", 14):
        ws.column_dimensions[col].width = w

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.55, bottom=0.55)
    ws.print_options.horizontalCentered = True

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
